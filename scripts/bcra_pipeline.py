#!/usr/bin/env python3
"""
Pipeline BCRA - Descarga y normaliza el archivo series.xlsm del Banco Central.

Produce en /data:
  bcra_hechos.csv   -> tabla de hechos en formato ANCHO (wide)
                       una fila por fecha, una columna por variable de nivel
                       Las variaciones se calculan con medidas DAX en Power BI.

Modelo Power BI:
  Tabla de hechos : bcra_hechos.csv  (fecha + 41 columnas numericas)
  Dimension unica : Calendario generado con CALENDARAUTO() en Power BI

Robustez ante cambios de esquema del BCRA:
  - Cambio de nombre de header    -> NO AFECTA (extraccion por letra de columna)
  - Cambio de fila de inicio      -> CUBIERTO  (deteccion dinamica de fila)
  - Adicion de nuevas columnas    -> DETECCION (se loguean en bcra_nuevas_columnas.json)
  - Serie particionada/reemplazada-> CUBIERTO  (validacion fuzzy del texto del header)
  - Si alguna validacion falla    -> PIPELINE FALLA con mensaje claro y GitHub notifica

Rezagos por hoja (verificados al 2026-04-19):
  TASAS DE MERCADO      -> diario,   ~3 dias rezago
  INSTRUMENTOS DEL BCRA -> diario,   ~3 dias rezago
  RESERVAS (saldos)     -> diario,   ~5 dias rezago
  BASE MONETARIA saldos -> diario,  ~18 dias rezago
  DEPOSITOS / PRESTAMOS -> mensual, ~49 dias rezago (fin de mes)
  Tasa politica monet.  -> discontinuada como serie separada desde ~jul 2025
"""

import csv
import io
import json
import os
import ssl
import sys
import urllib.request
from datetime import datetime, timezone

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

BCRA_URL = (
    "https://www.bcra.gob.ar/archivos/Pdfs/PublicacionesEstadisticas/series.xlsm"
)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.bcra.gob.ar/datos-monetarios-diarios/",
    "Accept": "application/vnd.ms-excel,*/*",
}

MAX_SCAN_ROWS = 30

# Minimo de palabras clave que deben aparecer en el header de la columna.
# 0.40 = al menos 40% de las palabras clave deben estar presentes.
HEADER_MATCH_THRESHOLD = 0.40

# (hoja, col, nombre_csv, descripcion, unidad, palabras_clave_header)
# Palabras clave: terminos sin tildes en minusculas que DEBEN aparecer en el
# texto del header de esa columna. Detectan si la columna fue reemplazada.
# Nota: el BCRA usa celdas combinadas con formulas; openpyxl en modo
# read_only devuelve formula-strings (empiezan con '=') que se filtran.
# [D]=diario  [M]=mensual (fin de mes)
COLUMNAS = [
    # RESERVAS [D] - millones USD
    ("RESERVAS", "C",  "res_total",             "Reservas internacionales - saldo total",        "millones USD", ["reservas", "stock", "total"]),
    ("RESERVAS", "D",  "res_oro_divisas",        "Reservas - oro, divisas y colocaciones",        "millones USD", ["oro", "divisas", "colocaciones"]),
    ("RESERVAS", "E",  "res_pase_pasivo",        "Reservas - pase pasivo USD exterior",           "millones USD", ["pase", "pasivo", "exterior"]),
    ("RESERVAS", "N",  "res_degs",               "DEGs 2009 - saldo asignaciones",                "millones USD", ["deg", "asignacion"]),
    ("RESERVAS", "P",  "tc_bcra",                "Tipo de cambio BCRA (pesos por USD)",           "pesos/USD",    ["tipo", "cambio"]),
    # BASE MONETARIA [D] - millones $
    ("BASE MONETARIA", "Z",  "bm_billetes_publico",   "BM - saldo billetes y monedas publico",   "millones $",   ["billetes", "publico"]),
    ("BASE MONETARIA", "AA", "bm_billetes_entidades", "BM - saldo billetes y monedas entidades", "millones $",   ["billetes", "entidades"]),
    ("BASE MONETARIA", "AC", "bm_cta_cte_bcra",       "BM - saldo cuentas corrientes en BCRA",  "millones $",   ["cuenta", "corriente", "bcra"]),
    # Col AD: celda combinada "Base Monetaria" esta en col Z; col AD solo
    # tiene "Total (12) = (8+9+10+11)". Keyword minimo: ["total"].
    ("BASE MONETARIA", "AD", "bm_saldo",               "BM - saldo base monetaria",              "millones $",   ["total"]),
    ("BASE MONETARIA", "AF", "bm_mas_cuasimonedas",    "BM - saldo base mas cuasimonedas",       "millones $",   ["total", "cuasimonedas"]),
    ("DEPOSITOS",      "AC", "m2",                     "M2 - agregado monetario amplio",          "millones $",   ["m2"]),
    ("DEPOSITOS",      "AD", "m2_transaccional_priv",  "M2 - transaccional privado",             "millones $",   ["m2", "transaccional"]),
    # DEPOSITOS [M] - millones
    ("DEPOSITOS", "K",  "dep_cc_priv",            "Depositos - CC sector privado",               "millones $",   ["cuenta", "corriente", "privado"]),
    ("DEPOSITOS", "L",  "dep_ca_priv",            "Depositos - CA sector privado",               "millones $",   ["caja", "ahorro", "privado"]),
    ("DEPOSITOS", "M",  "dep_pf_no_ajust_priv",   "Depositos - PF no ajustable privado",         "millones $",   ["plazo", "fijo", "ajustable"]),
    ("DEPOSITOS", "N",  "dep_pf_ajust_cer_priv",  "Depositos - PF ajustable CER/UVA privado",    "millones $",   ["cer", "uva", "ajustable"]),
    ("DEPOSITOS", "S",  "dep_total_pesos_priv",   "Depositos - total pesos sector privado",      "millones $",   ["total", "privado"]),
    ("DEPOSITOS", "U",  "dep_usd_priv_enpesos",   "Depositos - USD privado en pesos",            "millones $",   ["privado"]),
    ("DEPOSITOS", "X",  "dep_total_priv_enpesos", "Depositos - total pesos y USD privado en $",  "millones $",   ["total", "privado"]),
    ("DEPOSITOS", "AA", "dep_usd_priv",           "Depositos - USD sector privado en USD",       "millones USD", ["dolares", "privado"]),
    # PRESTAMOS [M] - millones
    ("PRESTAMOS", "B",  "prest_adelantos",        "Prestamos - adelantos CC pesos",              "millones $",   ["adelantos"]),
    ("PRESTAMOS", "C",  "prest_documentos",       "Prestamos - documentos pesos",                "millones $",   ["documentos"]),
    ("PRESTAMOS", "D",  "prest_hipotecarios",     "Prestamos - hipotecarios pesos",              "millones $",   ["hipotecarios"]),
    ("PRESTAMOS", "E",  "prest_prendarios",       "Prestamos - prendarios pesos",                "millones $",   ["prendarios"]),
    ("PRESTAMOS", "F",  "prest_personales",       "Prestamos - personales pesos",                "millones $",   ["personales"]),
    ("PRESTAMOS", "G",  "prest_tarjetas",         "Prestamos - tarjetas credito pesos",          "millones $",   ["tarjetas"]),
    ("PRESTAMOS", "H",  "prest_otros",            "Prestamos - otros pesos",                     "millones $",   ["otros"]),
    ("PRESTAMOS", "I",  "prest_total_pesos",      "Prestamos - total pesos",                     "millones $",   ["total", "pesos"]),
    ("PRESTAMOS", "Q",  "prest_total_usd",        "Prestamos - total USD en USD",                "millones USD", ["total", "dolares"]),
    ("PRESTAMOS", "U",  "prest_total_enpesos",    "Prestamos - total pesos y USD en pesos",      "millones $",   ["total", "privado"]),
    # TASAS DE MERCADO [D] - %
    ("TASAS DE MERCADO", "B",  "tasa_pf_total_tna",    "Tasa PF total general TNA",              "%",            ["plazo", "fijo", "general"]),
    ("TASAS DE MERCADO", "C",  "tasa_pf_personas_tna", "Tasa PF personas humanas TNA",           "%",            ["personas", "humanas"]),
    ("TASAS DE MERCADO", "L",  "badlar_total_tna",     "BADLAR total bancos TNA",                "%",            ["badlar", "total"]),
    ("TASAS DE MERCADO", "M",  "badlar_privados_tna",  "BADLAR bancos privados TNA",             "%",            ["badlar", "privados"]),
    ("TASAS DE MERCADO", "R",  "tasa_personales_tna",  "Tasa prestamos personales TNA",          "%",            ["personales"]),
    ("TASAS DE MERCADO", "S",  "tasa_adelantos_tna",   "Tasa adelantos empresas TNA",            "%",            ["adelantos"]),
    # INSTRUMENTOS DEL BCRA [D]
    ("INSTRUMENTOS DEL BCRA", "K",  "tasa_politica_tna", "Tasa politica monetaria TNA",         "%",            ["politica", "monetaria", "tna"]),
    ("INSTRUMENTOS DEL BCRA", "L",  "tasa_politica_tea", "Tasa politica monetaria TEA",         "%",            ["tea"]),
    ("INSTRUMENTOS DEL BCRA", "B",  "pases_pasivos",      "Pases pasivos - saldo total",        "millones $",   ["pases", "pasivos"]),
    ("INSTRUMENTOS DEL BCRA", "F",  "leliq_notaliq",      "LELIQ y NOTALIQ - saldo",            "millones $",   ["leliq", "notaliq"]),
    ("INSTRUMENTOS DEL BCRA", "AV", "lefi_entidades",     "LEFI - cartera entidades",           "millones $",   ["lefi", "letras", "fiscales"]),
]

# Columnas que ya estamos trackeando, por hoja (para detectar nuevas)
COLUMNAS_TRACKEADAS = {
    hoja: {col.strip() for (h, col, *_) in COLUMNAS if h == hoja}
    for hoja in {c[0] for c in COLUMNAS}
}

# Hojas de datos (excluir hojas de metadata)
HOJAS_DE_DATOS = {"BASE MONETARIA", "RESERVAS", "DEPOSITOS", "PRESTAMOS",
                   "TASAS DE MERCADO", "INSTRUMENTOS DEL BCRA"}


# ---------------------------------------------------------------------------
# FUNCIONES
# ---------------------------------------------------------------------------

def descargar_xlsx(url: str) -> bytes:
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=120, context=ctx) as resp:
        contenido = resp.read()
    if contenido[:2] != b"PK":
        raise ValueError(
            f"Respuesta inesperada del servidor (no es ZIP/Excel). "
            f"Primeros bytes: {contenido[:50]}"
        )
    return contenido


def detectar_fila_inicio(ws, max_scan: int = MAX_SCAN_ROWS) -> int:
    """
    Detecta la primera fila donde columna A contiene un datetime.
    Resiste si el BCRA agrega/quita filas de metadata arriba de los datos.
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        if isinstance(row[0], datetime):
            return i
    raise ValueError(
        f"No se encontro ninguna fila con fecha en las primeras {max_scan} filas "
        f"de la hoja '{ws.title}'. El BCRA puede haber cambiado la estructura."
    )


def leer_textos_header(ws, col_idx: int, max_scan: int = MAX_SCAN_ROWS) -> str:
    """
    Lee celdas de header de una columna. Descarta formula-strings (empiezan
    con '=') que openpyxl devuelve en modo read_only para celdas con formulas
    de concatenacion (patron del BCRA para construir labels de columnas).
    """
    textos = []
    for row in ws.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        celda = row[col_idx - 1]
        if isinstance(celda, str) and celda.strip() and not celda.strip().startswith("="):
            textos.append(celda)
    texto = " ".join(textos).lower()
    for c, s in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ü","u")]:
        texto = texto.replace(c, s)
    return texto


def validar_header(texto_header: str, palabras_clave: list, nombre_col: str) -> str | None:
    """
    Valida que al menos HEADER_MATCH_THRESHOLD de las palabras clave esten
    presentes en el header. Devuelve None si OK, mensaje de error si falla.
    """
    if not palabras_clave:
        return None
    encontradas = [p for p in palabras_clave if p in texto_header]
    ratio = len(encontradas) / len(palabras_clave)
    if ratio < HEADER_MATCH_THRESHOLD:
        return (
            f"[CRITICO] '{nombre_col}': el header de la columna no coincide "
            f"con lo esperado.\n"
            f"  Palabras esperadas : {palabras_clave}\n"
            f"  Palabras halladas  : {encontradas} ({ratio:.0%})\n"
            f"  Header actual      : {texto_header[:200]!r}\n"
            f"  ACCION: verificar si el BCRA movio o reemplazo esta serie."
        )
    return None


def detectar_columnas_nuevas(wb, fila_inicio_cache: dict) -> dict:
    """
    Escanea las hojas de datos en busca de columnas con datos numericos
    que NO esten en COLUMNAS_TRACKEADAS. Las loguea como 'disponibles
    pero no trackeadas' para que el usuario decida si agregarlas.

    Retorna dict: {hoja: [lista de letras de columnas nuevas]}
    """
    nuevas = {}
    for hoja in HOJAS_DE_DATOS:
        if hoja not in wb.sheetnames:
            continue
        ws = wb[hoja]
        row_start = fila_inicio_cache.get(hoja)
        if row_start is None:
            continue

        trackeadas = COLUMNAS_TRACKEADAS.get(hoja, set())
        columnas_con_datos = set()

        # Escanear solo las primeras 200 filas de datos para eficiencia
        for row in ws.iter_rows(min_row=row_start, max_row=row_start + 200, values_only=True):
            if not isinstance(row[0], datetime):
                continue
            for idx, valor in enumerate(row[1:], 2):  # col B en adelante
                if isinstance(valor, (int, float)):
                    columnas_con_datos.add(get_column_letter(idx))

        nuevas_en_hoja = columnas_con_datos - trackeadas
        if nuevas_en_hoja:
            nuevas[hoja] = sorted(nuevas_en_hoja)

    return nuevas


def extraer_serie(ws, col_letra: str, row_start: int) -> dict:
    col_idx = column_index_from_string(col_letra.strip())
    datos = {}
    for row in ws.iter_rows(min_row=row_start, values_only=True):
        fecha_raw = row[0]
        valor_raw = row[col_idx - 1]
        if not isinstance(fecha_raw, datetime):
            continue
        if valor_raw is None or not isinstance(valor_raw, (int, float)):
            continue
        datos[fecha_raw.strftime("%Y-%m-%d")] = float(valor_raw)
    return datos


# ---------------------------------------------------------------------------
# PIPELINE PRINCIPAL
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 60)
    print("Pipeline BCRA")
    print("=" * 60)

    print(f"\nDescargando: {BCRA_URL}")
    try:
        contenido = descargar_xlsx(BCRA_URL)
    except Exception as exc:
        print(f"\nERROR al descargar: {exc}", file=sys.stderr)
        sys.exit(1)
    print(f"Descarga OK: {len(contenido) / 1_000_000:.1f} MB")

    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, keep_vba=False)
    print(f"Hojas: {wb.sheetnames}\n")

    fila_inicio_cache: dict = {}
    series: dict = {}
    todas_las_fechas: set = set()
    errores: list = []
    ultima_fecha_por_col: dict = {}

    for (hoja, col, nombre_col, descripcion, unidad, palabras_clave) in COLUMNAS:
        if hoja not in wb.sheetnames:
            errores.append(f"[CRITICO] Hoja '{hoja}' no encontrada (columna '{nombre_col}')")
            series[nombre_col] = {}
            continue

        ws = wb[hoja]
        col_idx = column_index_from_string(col.strip())

        # Deteccion dinamica de fila de inicio (una vez por hoja)
        if hoja not in fila_inicio_cache:
            try:
                fila_inicio_cache[hoja] = detectar_fila_inicio(ws)
            except ValueError as exc:
                errores.append(str(exc))
                series[nombre_col] = {}
                continue
        row_start = fila_inicio_cache[hoja]

        # Validacion de header
        texto_header = leer_textos_header(ws, col_idx)
        error_header = validar_header(texto_header, palabras_clave, nombre_col)
        if error_header:
            errores.append(error_header)
            series[nombre_col] = {}
            continue

        # Extraccion de datos
        try:
            datos = extraer_serie(ws, col, row_start)
        except Exception as exc:
            errores.append(f"[CRITICO] Error extrayendo '{nombre_col}' (hoja={hoja}, col={col}): {exc}")
            series[nombre_col] = {}
            continue

        if not datos:
            errores.append(f"[AVISO] '{nombre_col}': 0 filas (hoja='{hoja}', col='{col}')")

        series[nombre_col] = datos
        todas_las_fechas.update(datos.keys())
        ultima = max(datos.keys()) if datos else "N/A"
        ultima_fecha_por_col[nombre_col] = ultima
        print(f"  {nombre_col:<30} | {len(datos):6,d} filas | hasta {ultima}")

    # Deteccion de columnas nuevas (no trackeadas)
    columnas_nuevas = detectar_columnas_nuevas(wb, fila_inicio_cache)
    wb.close()

    if columnas_nuevas:
        print("\nNUEVAS COLUMNAS DETECTADAS (disponibles pero no trackeadas):")
        for hoja, cols in columnas_nuevas.items():
            print(f"  Hoja '{hoja}': {cols}")
        print("  -> Revisar si alguna es relevante para agregar al pipeline.")

    nombres_col = [c[2] for c in COLUMNAS]
    fechas_ordenadas = sorted(todas_las_fechas)
    print(f"\nFechas unicas: {len(fechas_ordenadas):,}  |  Columnas: {len(nombres_col)}")

    output_dir = "data"
    os.makedirs(output_dir, exist_ok=True)

    # Tabla de hechos (wide format)
    with open(os.path.join(output_dir, "bcra_hechos.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["fecha"] + nombres_col)
        for fecha in fechas_ordenadas:
            fila = [fecha] + [series[col].get(fecha, "") for col in nombres_col]
            writer.writerow(fila)

    # Metadata de columnas (referencia; no se importa al modelo Power BI)
    with open(os.path.join(output_dir, "bcra_columnas.csv"), "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["columna", "descripcion", "unidad", "ultima_fecha"])
        for (hoja, col, nombre_col, descripcion, unidad, _) in COLUMNAS:
            writer.writerow([nombre_col, descripcion, unidad, ultima_fecha_por_col.get(nombre_col, "")])

    # JSON de control de actualizacion
    resumen = {
        "pipeline_ejecutado_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "total_fechas": len(fechas_ordenadas),
        "primera_fecha": fechas_ordenadas[0] if fechas_ordenadas else None,
        "ultima_fecha_global": fechas_ordenadas[-1] if fechas_ordenadas else None,
        "ultima_fecha_por_columna": ultima_fecha_por_col,
        "columnas_nuevas_detectadas": columnas_nuevas,
    }
    with open(os.path.join(output_dir, "bcra_last_update.json"), "w", encoding="utf-8") as f:
        json.dump(resumen, f, ensure_ascii=False, indent=2)

    print(f"\nbcra_hechos.csv  : {len(fechas_ordenadas):,} filas x {len(nombres_col) + 1} columnas")
    print(f"bcra_columnas.csv: {len(COLUMNAS)} columnas documentadas")
    print(f"bcra_last_update.json: generado")

    if errores:
        print("\n" + "=" * 60)
        print("ERRORES - INTERVENCION REQUERIDA:")
        for e in errores:
            print(f"\n{e}", file=sys.stderr)
        sys.exit(1)

    print("\nPipeline BCRA completado exitosamente.")


if __name__ == "__main__":
    main()
