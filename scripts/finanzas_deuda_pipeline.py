#!/usr/bin/env python3
"""
finanzas_deuda_pipeline.py
Descarga y normaliza el stock de deuda pública de la Secretaría de Finanzas.

Fuente: Boletín mensual de deuda bruta de la Administración Central.
  Página: https://www.argentina.gob.ar/economia/finanzas/datos-mensuales-de-la-deuda/datos
  Archivo: boletin_mensual_DD_MM_YYYY_N.xlsx (URL se detecta dinámicamente)
  Cobertura: 2019-01 en adelante, lag ~45-60 días.

Estructura del Excel (transpuesto: fechas en columnas, series en filas):
  Hoja A.1 - Deuda bruta total y componentes por instrumento/plazo
  Hoja A.3 - Composición por moneda y tasa

Salida:
  - data/finanzas_deuda_hechos.csv        wide format: fecha + 18 columnas
  - data/finanzas_deuda_last_update.json
"""

import io
import re
import ssl
import sys
import unicodedata
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from utils import (
    descargar_archivo, escribir_csv, escribir_json,
    log, log_error, timestamp_utc,
    BROWSER_HEADERS,
)

# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

PAGINA_DATOS = "https://www.argentina.gob.ar/economia/finanzas/datos-mensuales-de-la-deuda/datos"

DATA_DIR  = Path(__file__).parent.parent / "data"
CSV_PATH  = DATA_DIR / "finanzas_deuda_hechos.csv"
JSON_PATH = DATA_DIR / "finanzas_deuda_last_update.json"

# Configuración de series a extraer.
# El Excel es TRANSPUESTO: las fechas están en columnas (fila ~9-11),
# las series están en filas (label en col B = col 2).
# Formato: (hoja, fila_esperada, texto_parcial_label, nombre_col, descripcion)
# fila_esperada: número de fila donde está el label en col B (valor observado).
#   El parser busca el texto también en fila±5 y luego en toda la hoja,
#   por lo que resiste desplazamientos menores de hasta 5 filas.
SERIES_CONFIG = [
    # ── Hoja A.1: Deuda bruta y componentes por instrumento ──────────────────
    ("A.1", 10,  "A- DEUDA BRUTA",
     "deuda_bruta_total",      "Deuda bruta total Adm. Central (mill USD)"),
    ("A.1", 15,  "I- DEUDA EN SITUACION",
     "deuda_pago_normal",      "Deuda en situacion de pago normal (mill USD)"),
    ("A.1", 19,  "TITULOS PUBLICOS",
     "deuda_titulos_mlp",      "Titulos publicos mediano/largo plazo (mill USD)"),
    ("A.1", 83,  "LETRAS DEL TESORO",
     "deuda_letras_mlp",       "Letras del Tesoro mediano/largo plazo (mill USD)"),
    ("A.1", 95,  "PRESTAMOS",
     "deuda_prestamos_mlp",    "Prestamos mediano/largo plazo (mill USD)"),
    ("A.1", 99,  "ORGANISMOS INTERNACIONALES",
     "deuda_org_int",          "Prestamos de organismos internacionales (mill USD)"),
    ("A.1", 107, "FMI",
     "deuda_fmi",              "Deuda con el FMI (mill USD)"),
    ("A.1", 110, "ORGANISMOS OFICIALES",
     "deuda_bilaterales",      "Deuda bilateral (Club Paris + otros, mill USD)"),
    ("A.1", 123, "ADELANTOS TRANSITORIOS BCRA - Extraordinarios",
     "deuda_adelantos_ext",    "Adelantos transitorios BCRA extraordinarios (mill USD)"),
    ("A.1", 127, "ADELANTOS TRANSITORIOS BCRA - Ordinarios",
     "deuda_adelantos_ord",    "Adelantos transitorios BCRA ordinarios (mill USD)"),
    ("A.1", 136, "TITULOS PUBLICOS",
     "deuda_titulos_cp",       "Titulos publicos corto plazo (mill USD)"),
    ("A.1", 138, "LETRAS DEL TESORO",
     "deuda_letras_cp",        "Letras del Tesoro corto plazo (mill USD)"),
    ("A.1", 155, "DEUDA EN SITUACION DE PAGO DIFERIDO",
     "deuda_diferido",         "Deuda en situacion de pago diferido (mill USD)"),
    ("A.1", 160, "DEUDA ELEGIBLE PENDIENTE",
     "deuda_reestructuracion", "Deuda elegible pendiente de reestructuracion (mill USD)"),
    # ── Hoja A.3: Composicion por moneda (seccion 'COMPOSICION POR MONEDA') ──
    ("A.3", 54,  "Pesos no ajustable",
     "deuda_pesos_no_cer",     "Deuda en pesos no ajustable por CER (mill USD)"),
    ("A.3", 55,  "Pesos ajustable",
     "deuda_pesos_cer",        "Deuda en pesos ajustable por CER (mill USD)"),
    ("A.3", 56,  "Dolares",
     "deuda_usd",              "Deuda en dolares (mill USD)"),
    ("A.3", 57,  "Euros",
     "deuda_eur",              "Deuda en euros (mill USD)"),
    ("A.3", 59,  "DEG",
     "deuda_deg",              "Deuda en DEG (FMI, mill USD)"),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalizar(texto: str) -> str:
    """Minúsculas + eliminar tildes + strip."""
    txt = texto.strip().lower()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    return txt


def obtener_url_xlsx() -> str:
    """
    Descarga la página de datos y extrae la URL del xlsx más reciente.
    El markup tiene el patrón: blank:#https://...boletin_mensual_...xlsx
    Usa urllib directamente porque descargar_archivo() rechaza respuestas HTML.
    """
    log("Buscando URL del boletín mensual de deuda...")
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(PAGINA_DATOS, headers=BROWSER_HEADERS)
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
            html = resp.read().decode("utf-8", errors="ignore")
    except Exception as exc:
        raise RuntimeError(f"No se pudo descargar la página de Finanzas: {exc}") from exc
    match = re.search(r'blank:#(https://[^\s)"\'>]+\.xlsx)', html, re.IGNORECASE)
    if not match:
        raise RuntimeError(
            "No se encontró ningún enlace .xlsx en la página de Finanzas.\n"
            f"URL: {PAGINA_DATOS}\n"
            "Posible cambio en la estructura de la página — revisar manualmente."
        )
    return match.group(1)


def encontrar_map_fechas(ws, max_scan: int = 15) -> dict[int, str]:
    """
    Busca la primera fila donde ≥5 columnas contienen datetime.
    Devuelve {col_idx: "YYYY-MM-01"}.
    """
    for r in range(1, max_scan + 1):
        col_map = {}
        for c in range(2, min(ws.max_column + 1, 120)):
            v = ws.cell(r, c).value
            if isinstance(v, datetime):
                col_map[c] = v.strftime("%Y-%m-%d")
        if len(col_map) >= 5:
            return col_map
    return {}


def buscar_fila_label(ws, texto_parcial: str, fila_esperada: int,
                      ventana: int = 5) -> int | None:
    """
    Busca en col B (col 2) la fila cuyo texto normalizado contiene texto_parcial.
    Primero busca en fila_esperada±ventana, luego en toda la hoja.
    """
    needle = normalizar(texto_parcial)

    # Búsqueda en ventana esperada primero (más rápido y preciso)
    rango_cercano = range(
        max(1, fila_esperada - ventana),
        min(ws.max_row + 1, fila_esperada + ventana + 1)
    )
    for r in rango_cercano:
        v = ws.cell(r, 2).value
        if v and needle in normalizar(str(v)):
            return r

    # Fallback: toda la hoja
    for r in range(1, ws.max_row + 1):
        if r in rango_cercano:
            continue  # ya chequeado
        v = ws.cell(r, 2).value
        if v and needle in normalizar(str(v)):
            return r

    return None


def extraer_fila(ws, fila: int, col_fecha_map: dict[int, str]) -> dict[str, float]:
    """Extrae {fecha: valor} para todos los índices de fecha en col_fecha_map."""
    datos = {}
    for col_idx, fecha in col_fecha_map.items():
        v = ws.cell(fila, col_idx).value
        if isinstance(v, (int, float)):
            datos[fecha] = float(v)
    return datos


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    errores: list = []

    # 1. Obtener URL actual del xlsx
    try:
        xlsx_url = obtener_url_xlsx()
        log(f"  URL: {xlsx_url}")
    except RuntimeError as e:
        log_error(str(e))
        sys.exit(1)

    # 2. Descargar
    log("Descargando boletín de deuda...")
    try:
        contenido = descargar_archivo(xlsx_url)
        log(f"  OK — {len(contenido):,} bytes")
    except RuntimeError as e:
        log_error(str(e))
        sys.exit(1)

    # 3. Abrir workbook
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception as e:
        log_error(f"No se pudo abrir el xlsx: {e}")
        sys.exit(1)

    # 4. Extraer series
    log("Extrayendo series...")

    # Cache de mapas de fechas por hoja (una sola pasada por hoja)
    mapas_fechas: dict[str, dict[int, str]] = {}

    # {col_nombre: {fecha: valor}}
    series_datos: dict[str, dict[str, float]] = {}
    todas_las_fechas: set[str] = set()

    for (hoja, fila_esp, label, col_nombre, desc) in SERIES_CONFIG:
        if hoja not in wb.sheetnames:
            errores.append(f"[CRITICO] Hoja '{hoja}' no encontrada para '{col_nombre}'")
            series_datos[col_nombre] = {}
            continue

        ws = wb[hoja]

        # Construir mapa de fechas (una vez por hoja)
        if hoja not in mapas_fechas:
            mf = encontrar_map_fechas(ws)
            if not mf:
                errores.append(f"[CRITICO] No se encontró fila de fechas en hoja '{hoja}'")
                mapas_fechas[hoja] = {}
            else:
                mapas_fechas[hoja] = mf
                log(f"  Hoja '{hoja}': {len(mf)} fechas detectadas "
                    f"({min(mf.values())} -> {max(mf.values())})")

        col_fecha_map = mapas_fechas.get(hoja, {})
        if not col_fecha_map:
            series_datos[col_nombre] = {}
            continue

        # Buscar fila con el label
        fila_real = buscar_fila_label(ws, label, fila_esp)
        if fila_real is None:
            errores.append(f"[AVISO] '{col_nombre}': label '{label}' no encontrado en hoja '{hoja}'")
            series_datos[col_nombre] = {}
            continue

        # Avisar si el label se encontró lejos de donde esperábamos
        if abs(fila_real - fila_esp) > 5:
            log(f"  [!] '{col_nombre}': label encontrado en fila {fila_real} "
                f"(se esperaba ~{fila_esp}) — posible cambio de estructura")

        datos = extraer_fila(ws, fila_real, col_fecha_map)
        series_datos[col_nombre] = datos
        todas_las_fechas.update(datos.keys())
        log(f"  {col_nombre:<28} | {len(datos):3d} valores | hasta {max(datos) if datos else 'N/A'}")

    wb.close()

    # 5. Construir wide format
    nombres_cols   = [s[3] for s in SERIES_CONFIG]
    todas_columnas = ["fecha"] + nombres_cols
    fechas_ord     = sorted(todas_las_fechas)

    filas = []
    for fecha in fechas_ord:
        fila: dict = {"fecha": fecha}
        for col_nombre in nombres_cols:
            val = series_datos.get(col_nombre, {}).get(fecha)
            if val is not None:
                fila[col_nombre] = val
        filas.append(fila)

    # 6. Escribir CSV
    n = escribir_csv(CSV_PATH, filas, todas_columnas)
    log(f"CSV escrito: {CSV_PATH.name} ({n} filas x {len(todas_columnas)} columnas)")

    # Validar cobertura
    for col_nombre in nombres_cols:
        con_dato = sum(1 for f in filas if f.get(col_nombre) is not None)
        if con_dato == 0:
            errores.append(f"Columna '{col_nombre}' sin ningun dato")

    # 7. Metadata
    metadata = {
        "pipeline":             "finanzas_deuda",
        "ultima_actualizacion": timestamp_utc(),
        "total_filas":          n,
        "total_columnas":       len(todas_columnas) - 1,
        "fecha_inicio":         fechas_ord[0]  if fechas_ord else None,
        "fecha_fin":            fechas_ord[-1] if fechas_ord else None,
        "fuente":               xlsx_url,
        "nota_cobertura":       "Serie desde 2019-01. Lag de publicacion: ~45-60 dias.",
        "series_config": [
            {"hoja": s[0], "fila_esperada": s[1], "columna": s[3], "descripcion": s[4]}
            for s in SERIES_CONFIG
        ],
        "errores": errores,
    }
    escribir_json(JSON_PATH, metadata)
    log(f"JSON escrito: {JSON_PATH.name}")

    if errores:
        log_error(f"{len(errores)} error(es):")
        for e in errores:
            log_error(f"  {e}")
        sys.exit(1)

    log("Pipeline deuda pública completado sin errores.")


if __name__ == "__main__":
    main()
